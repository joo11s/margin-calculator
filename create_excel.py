#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""마진계산기.xlsx 생성 스크립트"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── 색상 정의 ──────────────────────────────────────────────
YELLOW   = PatternFill("solid", fgColor="FFFF00")
GREEN    = PatternFill("solid", fgColor="92D050")
ORANGE   = PatternFill("solid", fgColor="FF6600")
BLUE     = PatternFill("solid", fgColor="00B0F0")
HEADER   = PatternFill("solid", fgColor="1F4E79")
WHITE    = "FFFFFF"
BLACK    = "000000"

FONT_BASE    = Font(name="Arial", size=10)
FONT_HEADER  = Font(name="Arial", size=10, bold=True, color=WHITE)
FONT_ORANGE  = Font(name="Arial", size=10, bold=True, color=WHITE)
FONT_BLUE    = Font(name="Arial", size=10, color=BLACK)

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PLATFORMS = ["스마트스토어", "쿠팡", "카페24", "토스쇼핑", "기타"]
COMM_RATES = [0.0563, 0.106, 0.035, 0.11, 0.05]


def cell_style(cell, fill=None, font=None, align="center", border=True, fmt=None):
    if fill:
        cell.fill = fill
    cell.font = font or FONT_BASE
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        cell.border = BORDER
    if fmt:
        cell.number_format = fmt


def set_header(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    cell_style(c, fill=HEADER, font=FONT_HEADER)
    return c


def set_label(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    cell_style(c, font=FONT_BASE, align="left")
    return c


def set_input(ws, row, col, value=None):
    c = ws.cell(row=row, column=col, value=value)
    cell_style(c, fill=YELLOW, font=FONT_BASE, fmt='#,##0')
    return c


def set_auto(ws, row, col, formula, fmt='#,##0'):
    c = ws.cell(row=row, column=col, value=formula)
    cell_style(c, fill=GREEN, font=FONT_BASE, fmt=fmt)
    return c


def set_orange(ws, row, col, formula, fmt='#,##0'):
    c = ws.cell(row=row, column=col, value=formula)
    cell_style(c, fill=ORANGE, font=FONT_ORANGE, fmt=fmt)
    return c


def set_blue(ws, row, col, formula, fmt='#,##0'):
    c = ws.cell(row=row, column=col, value=formula)
    cell_style(c, fill=BLUE, font=FONT_BLUE, fmt=fmt)
    return c


# ═══════════════════════════════════════════════════════════
#  시트1: 마진계산기
# ═══════════════════════════════════════════════════════════
def build_calculator(ws):
    ws.title = "마진계산기"

    # ── 열 너비 ────────────────────────────────────────────
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    for col in "CDEFG":
        ws.column_dimensions[col].width = 14

    # ══ 섹션1: 원가 입력 (행 1~14) ═════════════════════════
    set_header(ws, 1, 1, "[ 섹션1 ] 원가 입력")
    ws.merge_cells("A1:B1")
    cell_style(ws["B1"], fill=HEADER, font=FONT_HEADER)

    labels_costs = [
        ("제조원가 / 공급가",  "B2"),
        ("포장재① (박스)",    "B3"),
        ("포장재② (완충재)",  "B4"),
        ("포장재③ (테이프)",  "B5"),
        ("포장재④ (라벨)",    "B6"),
        ("포장재⑤ (설명서)", "B7"),
        ("포장재⑥ (기타)",    "B8"),
        ("인건비",            "B9"),
        ("기타비용",          "B10"),
    ]
    for i, (lbl, _) in enumerate(labels_costs, start=2):
        set_label(ws, i, 1, lbl)

    # 입력 셀 (노란)
    for row in range(2, 11):
        set_input(ws, row, 2, 0)

    # 포장재 합계 (자동)
    set_label(ws, 11, 1, "포장재 합계")
    set_auto(ws, 11, 2, "=IFERROR(SUM(B3:B8),0)")

    # 총원가 합계
    set_label(ws, 12, 1, "총원가 합계")
    set_auto(ws, 12, 2, "=IFERROR(B2+B11+B9+B10,0)")

    # 과세 구분 (드롭다운)
    set_label(ws, 13, 1, "과세 구분")
    dv = DataValidation(type="list", formula1='"과세,면세"', allow_blank=False)
    ws.add_data_validation(dv)
    c_tax = ws.cell(row=13, column=2, value="과세")
    cell_style(c_tax, fill=YELLOW, font=FONT_BASE, fmt="@")
    dv.add(c_tax)

    # 부가세 (판매가는 아직 미결정 → 나중에 플랫폼별로 각각 계산)
    # 섹션3에서 판매가를 이용하므로 여기는 메모용 레이블만
    set_label(ws, 14, 1, "부가세 (※ 섹션3 자동계산)")
    c_vat_note = ws.cell(row=14, column=2, value="=IFERROR(IF(B13=\"과세\",\"판매가/11 적용\",\"면세\"),\"\")")
    cell_style(c_vat_note, fill=GREEN, font=FONT_BASE, fmt="@")

    # ══ 섹션2: 배송비 (행 16~19) ══════════════════════════
    set_header(ws, 16, 1, "[ 섹션2 ] 배송비")
    ws.merge_cells("A16:B16")
    cell_style(ws["B16"], fill=HEADER, font=FONT_HEADER)

    set_label(ws, 17, 1, "무료배송 기준금액")
    set_input(ws, 17, 2, 30000)

    set_label(ws, 18, 1, "배송비 단가")
    set_input(ws, 18, 2, 3000)

    # ── 섹션3 헤더 (행 20~) ───────────────────────────────
    sec3_start = 20

    set_header(ws, sec3_start, 1, "[ 섹션3 ] 플랫폼별 마진")
    ws.merge_cells(f"A{sec3_start}:G{sec3_start}")
    for col in range(2, 8):
        cell_style(ws.cell(row=sec3_start, column=col), fill=HEADER, font=FONT_HEADER)

    # 플랫폼 헤더 행
    ph = sec3_start + 1
    set_header(ws, ph, 1, "항목 \\ 플랫폼")
    for i, platform in enumerate(PLATFORMS):
        set_header(ws, ph, i + 2, platform)

    # 수수료율
    r_comm = ph + 1
    set_label(ws, r_comm, 1, "수수료율")
    for i, rate in enumerate(COMM_RATES):
        c = ws.cell(row=r_comm, column=i + 2, value=rate)
        cell_style(c, fill=GREEN, font=FONT_BASE, fmt="0.00%")

    # 판매가 입력 (노란)
    r_price = r_comm + 1
    set_label(ws, r_price, 1, "판매가 입력")
    for i in range(5):
        set_input(ws, r_price, i + 2, 0)

    # 배송비 부담
    r_ship = r_price + 1
    set_label(ws, r_ship, 1, "배송비 부담")
    price_col = {0: "B", 1: "C", 2: "D", 3: "E", 4: "F"}
    for i in range(5):
        pc = price_col[i]
        fml = f"=IFERROR(IF({pc}{r_price}>=$B$17,$B$18,0),0)"
        set_auto(ws, r_ship, i + 2, fml)

    # 수수료
    r_fee = r_ship + 1
    set_label(ws, r_fee, 1, "수수료")
    for i in range(5):
        pc = price_col[i]
        comm_cell = f"{price_col[i]}{r_comm}"
        fml = f"=IFERROR({pc}{r_price}*{comm_cell},0)"
        set_auto(ws, r_fee, i + 2, fml)

    # 부가세 (플랫폼별)
    r_vat = r_fee + 1
    set_label(ws, r_vat, 1, "부가세")
    for i in range(5):
        pc = price_col[i]
        fml = f"=IFERROR(IF($B$13=\"과세\",{pc}{r_price}/11,0),0)"
        set_auto(ws, r_vat, i + 2, fml)

    # 광고비 건당 (노란, 플랫폼별 개별)
    r_ad = r_vat + 1
    set_label(ws, r_ad, 1, "광고비 건당")
    for i in range(5):
        set_input(ws, r_ad, i + 2, 0)

    # 총비용
    r_total = r_ad + 1
    set_label(ws, r_total, 1, "총비용")
    for i in range(5):
        pc = price_col[i]
        fml = (f"=IFERROR($B$12+{pc}{r_ship}+{pc}{r_fee}"
               f"+{pc}{r_vat}+{pc}{r_ad},0)")
        set_auto(ws, r_total, i + 2, fml)

    # 마진 (주황, 흰글자)
    r_margin = r_total + 1
    set_label(ws, r_margin, 1, "마진")
    for i in range(5):
        pc = price_col[i]
        fml = f"=IFERROR({pc}{r_price}-{pc}{r_total},0)"
        set_orange(ws, r_margin, i + 2, fml)

    # 마진율 (주황)
    r_mrate = r_margin + 1
    set_label(ws, r_mrate, 1, "마진율")
    for i in range(5):
        pc = price_col[i]
        fml = f"=IFERROR({pc}{r_margin}/{pc}{r_price},0)"
        set_orange(ws, r_mrate, i + 2, fml, fmt="0.00%")

    # ══ 섹션4: 역산 (r_mrate+2~) ════════════════════════
    r_sec4 = r_mrate + 2
    set_header(ws, r_sec4, 1, "[ 섹션4 ] 역산 (목표마진율 기준 최소 판매가)")
    ws.merge_cells(f"A{r_sec4}:G{r_sec4}")
    for col in range(2, 8):
        cell_style(ws.cell(row=r_sec4, column=col), fill=HEADER, font=FONT_HEADER)

    ph4 = r_sec4 + 1
    set_header(ws, ph4, 1, "항목 \\ 플랫폼")
    for i, platform in enumerate(PLATFORMS):
        set_header(ws, ph4, i + 2, platform)

    # 목표마진율 (노란)
    r_target = ph4 + 1
    set_label(ws, r_target, 1, "목표 마진율")
    for i in range(5):
        c = ws.cell(row=r_target, column=i + 2, value=0.3)
        cell_style(c, fill=YELLOW, font=FONT_BASE, fmt="0.00%")

    # 최소판매가 (파란)
    r_minprice = r_target + 1
    set_label(ws, r_minprice, 1, "최소 판매가")
    for i in range(5):
        pc = price_col[i]
        comm_cell = f"{pc}{r_comm}"
        target_cell = f"{pc}{r_target}"
        fml = (f"=IFERROR(CEILING(($B$12+$B$18)"
               f"/(1-{comm_cell}-{target_cell}),100),0)")
        set_blue(ws, r_minprice, i + 2, fml)

    # ══ 섹션5: BEP ═══════════════════════════════════════
    r_sec5 = r_minprice + 2
    set_header(ws, r_sec5, 1, "[ 섹션5 ] BEP (손익분기점)")
    ws.merge_cells(f"A{r_sec5}:G{r_sec5}")
    for col in range(2, 8):
        cell_style(ws.cell(row=r_sec5, column=col), fill=HEADER, font=FONT_HEADER)

    ph5 = r_sec5 + 1
    set_header(ws, ph5, 1, "항목 \\ 플랫폼")
    for i, platform in enumerate(PLATFORMS):
        set_header(ws, ph5, i + 2, platform)

    # 월고정비 (노란)
    r_fixed = ph5 + 1
    set_label(ws, r_fixed, 1, "월 고정비")
    for i in range(5):
        set_input(ws, r_fixed, i + 2, 0)

    # BEP 수량 (파란)
    r_bep_qty = r_fixed + 1
    set_label(ws, r_bep_qty, 1, "BEP 수량")
    for i in range(5):
        pc = price_col[i]
        margin_cell = f"{pc}{r_margin}"
        fixed_cell  = f"{pc}{r_fixed}"
        fml = f"=IFERROR(CEILING({fixed_cell}/{margin_cell},1),0)"
        set_blue(ws, r_bep_qty, i + 2, fml)

    # BEP 매출 (파란)
    r_bep_rev = r_bep_qty + 1
    set_label(ws, r_bep_rev, 1, "BEP 매출")
    for i in range(5):
        pc = price_col[i]
        fml = f"=IFERROR({pc}{r_bep_qty}*{pc}{r_price},0)"
        set_blue(ws, r_bep_rev, i + 2, fml)

    # 행 높이
    for row in range(1, r_bep_rev + 1):
        ws.row_dimensions[row].height = 18

    return ws


# ═══════════════════════════════════════════════════════════
#  시트2: 수수료 기준표
# ═══════════════════════════════════════════════════════════
def build_fee_table(ws):
    ws.title = "수수료기준표"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 22

    headers = ["플랫폼", "서비스", "수수료율", "비고", "업데이트"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        cell_style(c, fill=HEADER, font=FONT_HEADER)

    data = [
        # 스마트스토어
        ("스마트스토어", "기본 판매 수수료", "5.63%",
         "매출 연동 수수료 (부가세 별도)\n카테고리에 따라 상이 (2~6%)", "2026.06"),
        ("스마트스토어", "결제 수수료", "포함",
         "판매 수수료에 포함", "2026.06"),
        # 쿠팡
        ("쿠팡", "로켓그로스 수수료", "10.6%",
         "풀필먼트+판매수수료 통합\n(카테고리별 8~12%)", "2026.06"),
        ("쿠팡", "마켓플레이스 수수료", "5.5~11%",
         "카테고리별 상이\n기본 적용률 10.6% 사용", "2026.06"),
        # 카페24
        ("카페24", "거래 수수료", "3.5%",
         "월정액 플랜에 따라 0~3.5%\n기본 스탠다드 기준", "2026.06"),
        ("카페24", "PG 결제 수수료", "별도",
         "카드사별 2.0~3.3% 별도 발생", "2026.06"),
        # 토스쇼핑
        ("토스쇼핑", "판매 수수료", "11.0%",
         "토스페이먼츠 결제 수수료 포함\n(카테고리별 9~13%)", "2026.06"),
        # 기타
        ("기타 (사용자 설정)", "수수료", "5.0%",
         "직접 수정 가능 (기본값 5%)", "2026.06"),
    ]

    fill_map = {
        "스마트스토어": PatternFill("solid", fgColor="E2EFDA"),
        "쿠팡":        PatternFill("solid", fgColor="FCE4D6"),
        "카페24":      PatternFill("solid", fgColor="DDEBF7"),
        "토스쇼핑":    PatternFill("solid", fgColor="EDE7F6"),
        "기타 (사용자 설정)": PatternFill("solid", fgColor="FFF9C4"),
    }

    for row_idx, (platform, service, rate, note, updated) in enumerate(data, start=2):
        fill = fill_map.get(platform, PatternFill("solid", fgColor="FFFFFF"))
        values = [platform, service, rate, note, updated]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill = fill
            c.font = FONT_BASE
            c.alignment = Alignment(
                horizontal="center", vertical="center",
                wrap_text=True
            )
            c.border = BORDER
        ws.row_dimensions[row_idx].height = 36

    # 주석 행
    note_row = len(data) + 3
    c = ws.cell(row=note_row, column=1,
                value="※ 위 수수료율은 2026년 6월 기준이며, 각 플랫폼 정책 변경 시 업데이트 필요합니다.")
    c.font = Font(name="Arial", size=9, italic=True, color="666666")
    ws.merge_cells(f"A{note_row}:E{note_row}")


# ═══════════════════════════════════════════════════════════
#  메인
# ═══════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    ws1 = wb.active
    build_calculator(ws1)

    ws2 = wb.create_sheet()
    build_fee_table(ws2)

    out = "마진계산기.xlsx"
    wb.save(out)
    print(f"✅  {out} 생성 완료")


if __name__ == "__main__":
    main()
